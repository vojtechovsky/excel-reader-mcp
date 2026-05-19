"""MCP stdio server: read Excel (.xlsx) via openpyxl."""

from __future__ import annotations

import asyncio
import json
import os
from datetime import date, datetime, time, timedelta
from decimal import Decimal

import mcp.server.stdio
import mcp.types as types
from mcp.server import NotificationOptions, Server
from mcp.server.models import InitializationOptions
from openpyxl import load_workbook

server = Server("excel-reader-server")


def _cell_value(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, time):
        return v.isoformat()
    if isinstance(v, timedelta):
        return str(v)
    if isinstance(v, Decimal):
        return float(v)
    return v


def _rows_matrix(ws) -> list[list]:
    out: list[list] = []
    for row in ws.iter_rows(values_only=True):
        out.append([_cell_value(c) for c in row])
    return out


@server.list_tools()
async def handle_list_tools() -> list[types.Tool]:
    return [
        types.Tool(
            name="read_excel",
            description="Read content from Excel (xlsx) files",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file",
                    },
                },
                "required": ["file_path"],
            },
        ),
        types.Tool(
            name="read_excel_by_sheet_name",
            description=(
                "Read content from a specific sheet by name in Excel (xlsx) files. "
                "Reads first sheet if sheet_name not provided."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet to read (optional, defaults to first sheet)",
                    },
                },
                "required": ["file_path"],
            },
        ),
        types.Tool(
            name="read_excel_by_sheet_index",
            description=(
                "Read content from a specific sheet by index in Excel (xlsx) files. "
                "Reads first sheet (index 0) if sheet_index not provided."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file",
                    },
                    "sheet_index": {
                        "type": "integer",
                        "description": "Index of the sheet to read (optional, defaults to 0)",
                        "minimum": 0,
                    },
                },
                "required": ["file_path"],
            },
        ),
    ]


@server.call_tool()
async def handle_call_tool(name: str, arguments: dict | None) -> list[types.TextContent]:
    if name not in ("read_excel", "read_excel_by_sheet_name", "read_excel_by_sheet_index"):
        raise ValueError(f"Unknown tool: {name}")
    if not arguments:
        raise ValueError("Missing arguments")

    file_path = arguments.get("file_path")
    if not file_path:
        raise ValueError("file_path is required")

    abs_path = os.path.abspath(file_path)
    if not os.path.isfile(abs_path):
        raise ValueError(f"File not found: {abs_path}")

    try:
        if name == "read_excel":
            with load_workbook(abs_path, read_only=True, data_only=True) as wb:
                sheets: dict[str, list[list]] = {}
                for sn in wb.sheetnames:
                    sheets[sn] = _rows_matrix(wb[sn])
            payload = {"status": "ok", "file_path": abs_path, "sheets": sheets}

        elif name == "read_excel_by_sheet_name":
            sheet_name = arguments.get("sheet_name")
            with load_workbook(abs_path, read_only=True, data_only=True) as wb:
                if sheet_name:
                    if sheet_name not in wb.sheetnames:
                        raise ValueError(f"Unknown sheet_name: {sheet_name!r}")
                    ws = wb[sheet_name]
                    used_name = sheet_name
                else:
                    used_name = wb.sheetnames[0]
                    ws = wb[used_name]
                rows = _rows_matrix(ws)
            payload = {
                "status": "ok",
                "file_path": abs_path,
                "sheet_name": used_name,
                "rows": rows,
            }

        else:  # read_excel_by_sheet_index
            idx = arguments.get("sheet_index", 0)
            if not isinstance(idx, int) or idx < 0:
                raise ValueError("sheet_index must be a non-negative integer")
            with load_workbook(abs_path, read_only=True, data_only=True) as wb:
                if idx >= len(wb.sheetnames):
                    raise ValueError(
                        f"sheet_index {idx} out of range (0..{len(wb.sheetnames) - 1})"
                    )
                used_name = wb.sheetnames[idx]
                rows = _rows_matrix(wb[used_name])
            payload = {
                "status": "ok",
                "file_path": abs_path,
                "sheet_index": idx,
                "sheet_name": used_name,
                "rows": rows,
            }

        return [types.TextContent(type="text", text=json.dumps(payload, ensure_ascii=False))]
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {e}") from e


async def _async_main() -> None:
    async with mcp.server.stdio.stdio_server() as (read_stream, write_stream):
        await server.run(
            read_stream,
            write_stream,
            InitializationOptions(
                server_name="excel-reader-server",
                server_version="0.1.0",
                capabilities=server.get_capabilities(
                    notification_options=NotificationOptions(),
                    experimental_capabilities={},
                ),
            ),
        )


def main() -> None:
    asyncio.run(_async_main())


if __name__ == "__main__":
    main()
